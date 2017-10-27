import os
import time
import json
from tkinter import filedialog
from tkinter import messagebox

import openpyxl
import requests

class NYK_Track(object):
    """Track for NYK shipping container"""

    def __init__(self):

        self.headers = {
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'en-US,en;q=0.8,zh-CN;q=0.6,zh;q=0.4,zh-TW;q=0.2',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Host': 'www.nykline.com',
            'X-Requested-With': 'XMLHttpRequest',
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.101 Safari/537.36',
            'Referer': 'https://www.nykline.com/ecom/CUP_HOM_3301.do?sessLocale=en',
        }
        # General requests Session object with headers
        self.session = requests.Session()
        self.session.headers.update(self.headers)

        self.url0 = "https://www.nykline.com/ecom/MenuGS.do?f_cmd=105&pagerows=&mnu_div_cd=E&hpg_lang_tp_cd=EN&user_id="
        self.url1 = "https://www.nykline.com/ecom/MenuGS.do?f_cmd=101&pagerows=&dsp_flg=&mnu_div_cd=F&hpg_lang_tp_cd=EN"
        self.url2 = "https://www.nykline.com/ecom/apps/gnoss/webservice/tracktrace/cargotracking/script/CUP_HOM_3301.js?version=1482308284000&_={timestamp}"
        self.url3 = "https://www.nykline.com/ecom/CUP_HOM_3301GS.do"

    
    def get_bkg_cop_no(self, cntr_no, timestamp):
        # Get the bkg number and cop number

        data1 = {
            'cntr_no': cntr_no,
            'cust_cd': '',
            'f_cmd': 122
        }
        
        self.session.get(self.url0)
        self.session.get(self.url1)
        self.session.get(self.url2.format(timestamp=timestamp))
        res = self.session.post(self.url3, data=data1)
        res_json = res.json()
        bkg_no = res_json['list'][0]['bkgNo']
        cop_no = res_json['list'][0]['copNo']
        return (bkg_no, cop_no)

    def track_and_record(self, cntr_no, bkg_no, cop_no):
        # Use the container number to search the logistics
        # The tracking result save in a list 
        # which every item inside it is a dict

        data2 = {
            'cntr_no': cntr_no,
            'bkg_no': bkg_no,
            'cop_no': cop_no,
            'f_cmd': 125
        }
        res = self.session.post(self.url3, data=data2)
        res_json = res.json()
        items = res_json['list']
        result = []
        for item in items:
            d = {}
            d['no'] = item['no']
            d['status'] = item['statusNm']
            d['place'] = item['placeNm']
            d['eventDt'] = item['eventDt']
            result.append(d)

        return result

    def write_to_excel(self, ws, ctn_row, cntr_no, result):
        # write the 6 key point within the result to NYK_Result sheet
        # 2、4、5、8、9、13、14
        ws.cell(row=ctn_row, column=2).value = result[2]['eventDt'].split(' ')[0]
        ws.cell(row=ctn_row, column=3).value = result[4]['eventDt'].split(' ')[0]
        ws.cell(row=ctn_row, column=4).value = result[5]['eventDt'].split(' ')[0]
        ws.cell(row=ctn_row, column=5).value = result[8]['eventDt'].split(' ')[0]
        ws.cell(row=ctn_row, column=6).value = result[9]['eventDt'].split(' ')[0]
        ws.cell(row=ctn_row, column=7).value = result[13]['eventDt'].split(' ')[0]
        ws.cell(row=ctn_row, column=8).value = result[14]['eventDt'].split(' ')[0]


def get_max_row(ws_obj):
    key = True
    i = 0
    while key:
        i = i + 1
        value = ws_obj.cell(row=i, column=1).value
        if value is None:
            key = False
            return i - 1


if __name__ == "__main__":

    # General unix timestamp
    timestamp = int(round(time.time() * 1000))

    #**************************************************** Get container number list ***************************************#
    #
    #
    # Get the Excel workbook and inner container worksheet
    try:
        with open("config.cfg") as f:
            config_json = json.loads(f.read().strip().replace("'", '"'))
    except FileNotFoundError:           # If the configuration file does not exists, show a warning
        messagebox.showwarning(title="NotFound", message="The config.cfg file not found in current directory")
    else:
        filename = config_json['filePath']

        # If the filename path is void or file not exists
        if filename == "" or not os.path.exists(filename):
            # filename = filedialog.askopenfilename(filetypes=("Excel", ["*.xlsx", "*.xls"]), initialdir="C:/")
            filename = filedialog.askopenfilename(initialdir="C:/")
            config_json['filePath'] = filename
            with open("config.cfg", "w") as f:
                f.write(str(config_json))

    wb = openpyxl.load_workbook(filename)
    # container_ws = wb.get_sheet_by_name("Container_no")
    nyk_ws = wb.get_sheet_by_name("Tracking_Result")
    # ctn_last_row = get_max_row(container_ws)
    result_last_row = get_max_row(nyk_ws)
    container_list = [nyk_ws.cell(row=r, column=1).value for r in range(2, result_last_row + 1)]


    #*************************************************** Start to crawl ****************************************************#
    nyk_track = NYK_Track()
    for ctn_row, ctn in enumerate(container_list):
        print("Tracking %s ......" % ctn, end="")
        try:
            bkg_no, cop_no = nyk_track.get_bkg_cop_no(ctn, timestamp)
        except KeyError:
            continue

        logistics_result = nyk_track.track_and_record(ctn, bkg_no, cop_no)
        print("Write to excel...", end="")
        nyk_track.write_to_excel(nyk_ws, ctn_row + 2, ctn, logistics_result)
        print("Done!")
        time.sleep(2)
    else:
        print("Finished all!")

    # Save the file
    wb.save(filename)
    wb.close()
    wb = None

    # Open the result excel file in Excel Application
    os.startfile(filename)