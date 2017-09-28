import os
import openpyxl
from lxml import etree

class HtmlProcess(object):
    """处理SAP.ZDER里面导出的htm文件，并把结果写入到对应的excel文件中"""

    def __init__(self):
        self.dc_list = ["A668", "A672", "A673", "A680", "A681", "A715", "A716", "B194"]

    def get_data(self, fname, code_list):
        """接受文件名和Code列表来处理数据"""
        
        src = open(fname).read()
        htm = etree.HTML(src)
        p = htm.xpath("//blockquote/p[2]")[0]
        trs = p.findall(".//tr")
        # 开头行是表头，结尾行是表尾，中间还有位于第2页的表头，是第41个tr
        # 这3行都不要
        trs = trs[1:41] + trs[42:50]

        dc_list = self.dc_list
        # 创建字典以存放库存数据
        result_dict = {}
        for d in dc_list:
            result_dict[d] = {}
            for c in code_list:
                result_dict[d][c] = {"dc": 0, "intransit": 0}

        for tr in trs:
            code = tr.find(".//td[3]/font/nobr").text.strip()                         # Material Code
            dc = tr.find(".//td[5]/font/nobr").text.strip()                           # DC
            unres_inv = int(float(tr.find(".//td[7]/font/nobr").text.strip().replace(",", "")))        # Unrestrict Inventory
            qi_inv = int(float(tr.find(".//td[9]/font/nobr").text.strip().replace(",", "")))           # Quality Inventory
            block_inv = int(float(tr.find(".//td[10]/font/nobr").text.strip().replace(",", "")))       # Block Inventory
            intransit_inv = int(float(tr.find(".//td[13]/font/nobr").text.strip().replace(",", "")))   # Intransit Inventory
            result_dict[dc][code]["dc"] = unres_inv + qi_inv + block_inv
            result_dict[dc][code]["intransit"] = intransit_inv

        return result_dict

    def get_ws(self, fname):
        wb = openpyxl.load_workbook(fname, keep_vba=True)
        ws = wb.get_sheet_by_name("Sheet2")
        return (wb, ws)

    def write_to_excel(self, ws, code_list, result_dict):
        '''把结果写入到ws中，ws是一个Excel的工作表对象'''
        dc_list = self.dc_list
        for col, code in zip(range(2, 14, 2), code_list):
            for r, dc in zip(range(8, 16), dc_list):
                ws.cell(row=r, column=col).value = result_dict[dc][code]["dc"]
                ws.cell(row=r, column=col+1).value = result_dict[dc][code]["intransit"]
            

if __name__ == "__main__":
    htmlProcess = HtmlProcess()

    # 处理koala数据
    koala_htm = "C:/users/xian.jr/Documents/SAP/SAP GUI/ka.htm"
    koala_inv = "C:/Users/xian.jr/Documents/PG/Jacqueline/Koala/Koala Inventory Template.xlsm"
    koala_code_list = ["82261753", "82261756", "82261762", "82261951", "82263780", "82264483"]
    koala_result_dict = htmlProcess.get_data(koala_htm, koala_code_list)
    koala_inv_wb, koala_inv_ws = htmlProcess.get_ws(koala_inv)
    htmlProcess.write_to_excel(koala_inv_ws, koala_code_list, koala_result_dict)
    koala_inv_wb.save(koala_inv)


    # 处理tampax数据
    # tampax_htm = "C:/users/xian.jr/Documents/SAP/SAP GUI/ta.htm"
    # tampax_inv = "C:/Users/xian.jr/Documents/PG/Jacqueline/Tampax/Inventory Template.xlsm"
    # tampax_code_list = ["82255634", "82258740", "82259122", "82258746", "82258747", "82261076"]
    # tampax_result_dict = htmlProcess.get_data(tampax_htm, tampax_code_list)
    # tampax_inv_wb, tampax_inv_ws = htmlProcess.get_ws(tampax_inv)
    # htmlProcess.write_to_excel(tampax_inv_ws, tampax_code_list, tampax_result_dict)
    # tampax_inv_wb.save(tampax_inv)